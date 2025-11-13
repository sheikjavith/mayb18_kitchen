
from flask import Flask, request, jsonify, send_file, render_template_string
from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import datetime
import os

app = Flask(__name__)

MENU_XLSX = Path("menu.xlsx")
BILLS_XLSX = Path("bills.xlsx")
TABLES = ["Outside 1", "Outside 2", "Swiggy", "Inside 1", "Inside 2", "Inside 3", "Last 1", "Last 2"]

# -----------------------------
# Excel Utilities
# -----------------------------
def ensure_menu_file():
    if not MENU_XLSX.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(["Category", "Item Name", "Price"])
        wb.save(MENU_XLSX)
        wb.close()

def ensure_bills_file():
    if not BILLS_XLSX.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(["Bill No", "Date & Time", "Item Name", "Qty", "Rate", "Amount", "Total", "Payment Method", "Table"])
        wb.save(BILLS_XLSX)
        wb.close()

def load_menu_from_xlsx():
    ensure_menu_file()
    wb = load_workbook(MENU_XLSX)
    sheet = wb.active
    data = {}
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if idx == 1:
            continue
        cells = list(row)[:3]
        if len(cells) < 2:
            continue
        cat = str(cells[0] or "").strip()
        name = str(cells[1] or "").strip()
        try:
            price = float(cells[2] or 0)
        except:
            continue
        if not name:
            continue
        data.setdefault(cat or "Uncategorized", []).append({"name": name, "price": price})
    wb.close()
    return data

def write_menu_to_xlsx(menu_dict):
    wb = Workbook()
    ws = wb.active
    ws.append(["Category", "Item Name", "Price"])
    for cat, items in menu_dict.items():
        for it in items:
            ws.append([cat, it["name"], it["price"]])
    wb.save(MENU_XLSX)
    wb.close()

def next_bill_no():
    ensure_bills_file()
    wb = load_workbook(BILLS_XLSX, read_only=True)
    sheet = wb.active
    max_bn = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            val = int(row[0])
            max_bn = max(max_bn, val)
        except:
            pass
    wb.close()
    return max_bn + 1

def append_bill_to_xlsx(bill_obj):
    ensure_bills_file()
    wb = load_workbook(BILLS_XLSX)
    sheet = wb.active
    bn = bill_obj.get("billNo", next_bill_no())
    raw_dt = bill_obj.get("dateTime", None)

    if isinstance(raw_dt, datetime):
        dt = raw_dt
    elif isinstance(raw_dt, str):
        try:
            dt = datetime.fromisoformat(raw_dt.replace("Z", "").split(".")[0])
        except:
            try:
                dt = datetime.strptime(raw_dt, "%Y-%m-%dT%H:%M:%S")
            except:
                dt = datetime.now()
    else:
        dt = datetime.now()

    if dt.tzinfo:
        dt = dt.replace(tzinfo=None)

    payment = bill_obj.get("payment", "")
    total = float(bill_obj.get("total", 0.0))
    table = bill_obj.get("table", "")

    for it in bill_obj.get("items", []):
        name = it.get("name", "")
        qty = int(it.get("qty", 0))
        rate = float(it.get("rate", 0))
        amount = float(it.get("amount", 0))
        sheet.append([bn, dt, name, qty, rate, amount, total, payment, table])
        sheet.cell(row=sheet.max_row, column=2).number_format = 'yyyy-mm-dd hh:mm AM/PM'

    wb.save(BILLS_XLSX)
    wb.close()

def read_bills_from_xlsx():
    ensure_bills_file()
    wb = load_workbook(BILLS_XLSX)
    sheet = wb.active
    bills = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or all(c is None for c in row):
            continue
        bill_no, dt, item_name, qty, rate, amount, total, payment, table = (list(row) + [None]*9)[:9]
        key = str(bill_no)
        if key not in bills:
            bills[key] = {
                "billNo": bill_no,
                "dateTime": dt,
                "table": table,
                "payment": payment,
                "total": total,
                "items": []
            }
        bills[key]["items"].append({
            "name": item_name, "qty": qty, "rate": rate, "amount": amount
        })
    wb.close()
    return list(bills.values())

# -----------------------------
# Flask Routes
# -----------------------------
@app.route("/")
def index():
    try:
        tpl = open("index.html", "r", encoding="utf-8").read()
    except FileNotFoundError:
        tpl = "<h2>Missing index.html</h2><p>Please place your HTML file here.</p>"
    return render_template_string(tpl, tables=TABLES)

@app.route("/api/menu", methods=["GET", "POST"])
def api_menu():
    if request.method == "GET":
        return jsonify(load_menu_from_xlsx())
    payload = request.get_json(force=True)
    cat = str(payload.get("category", "")).strip()
    name = str(payload.get("name", "")).strip()
    try:
        price = float(payload.get("price", 0))
    except:
        price = 0.0
    if not name:
        return ("Missing name", 400)
    menu = load_menu_from_xlsx()
    menu.setdefault(cat or "Uncategorized", []).append({"name": name, "price": price})
    write_menu_to_xlsx(menu)
    return ("OK", 200)

@app.route("/api/next_bill_no")
def api_next_bill():
    return jsonify({"next": next_bill_no()})

@app.route("/api/bills", methods=["GET", "POST"])
def api_bills():
    if request.method == "GET":
        bills = read_bills_from_xlsx()

        # ✅ Filter only today's bills (00:00 to 23:59)
        now = datetime.now()
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end = now.replace(hour=23, minute=59, second=59, microsecond=999999)

        filtered = []
        for b in bills:
            dt = b.get("dateTime")
            try:
                if isinstance(dt, str):
                    cleaned = dt.replace("Z", "").split("+")[0].split(".")[0]
                    dt = datetime.fromisoformat(cleaned)
                if isinstance(dt, datetime) and start <= dt <= end:
                    filtered.append(b)
            except Exception as e:
                print("Date parse error:", e)
        return jsonify(filtered)

    else:
        data = request.get_json(force=True)
        if not isinstance(data.get("items", []), list):
            return ("Invalid items", 400)
        if not data.get("billNo"):
            data["billNo"] = next_bill_no()
        if not data.get("dateTime"):
            data["dateTime"] = datetime.now().isoformat()
        append_bill_to_xlsx(data)
        return ("OK", 200)

@app.route("/download/<fname>")
def download(fname):
    if fname not in ("menu.xlsx", "bills.xlsx"):
        return ("Forbidden", 403)
    if not Path(fname).exists():
        if fname == "menu.xlsx":
            ensure_menu_file()
        else:
            ensure_bills_file()
    return send_file(fname, as_attachment=True)

if __name__ == "__main__":
    ensure_menu_file()
    ensure_bills_file()
    if not Path("index.html").exists():
        print("⚠️ Missing index.html — please place it in the same folder.")
    app.run(host="127.0.0.1", port=5000, debug=True)